import openpyxl
import os
from datetime import time
from datetime import datetime, timedelta, date
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill, Font, Border, Side
import pandas as pd
import logging
import glob
import tkinter as tk
from tkinter import filedialog

import logging
logging.basicConfig(level=logging.DEBUG, filename="script_log.log", filemode="w", format='%(levelname)s - %(message)s')

 

 

 
def calculate_breaks(shift_start, shift_end):
    duration = time_difference(shift_start, shift_end)
    total_hours = duration.total_seconds() / 3600
    if total_hours <= 5:
        return 10
    else:
        return "10 | 30 | 10"

 

def parse_time(value):
    hours, minutes = map(int, value.split(":"))
    return time(hours, minutes)

 

def extract_role_and_time(cell_value):
    lines = cell_value.split("\n")
    roles_and_times = []
    # Loop through lines two at a time (time and role)
    
    for i in range(0, len(lines), 2):
        time_value = lines[i].strip()
        role_value = lines[i + 1].strip() if i + 1 < len(lines) else None
       
        # Check if time and role values are valid
        if time_value and role_value:
            roles_and_times.append((abbreviate_role(role_value), time_value))
   
    return roles_and_times

 

 

def time_difference(start, end):
    start_dt = datetime.combine(datetime.today(), start)
    end_dt = datetime.combine(datetime.today(), end)
    if end < start:
        end_dt += timedelta(days=1)
    return end_dt - start_dt

 

 

def get_initials(name):
    if not name:
        return ""
    parts = name.split()
    initials = "".join([p[0] for p in parts])
    return initials

 

def get_date_from_day_name(day_name):
    today = date.today()
    days_mapping = {
        "Monday": 0,
        "Tuesday": 1,
        "Wednesday": 2,
        "Thursday": 3,
        "Friday": 4,
        "Saturday": 5,
        "Sunday": 6
    }

   

    # Get the difference in days from today to the desired day
    day_difference = days_mapping[day_name] - today.weekday()
    if day_difference < 0:
        day_difference += 7

 

    # Calculate the new date
    target_date = today + timedelta(days=day_difference)

    return target_date.strftime("%A, %d %b %Y")

 

def sort_key(row):
    role = row[0]
    start_time = row[2] if len(row) > 2 else ""

   

    if role == 'DM':
        role_priority = 2
    elif role == 'CL':
        role_priority = 3
    elif role == 'T':
        role_priority = 4
    elif role == 'PS':
        role_priority = 5
    else:
        role_priority = 1
    return (role_priority, start_time)

  


def abbreviate_role(role):
    role_mapping = {
        'Lifeguard': 'LG',
        'Duty Manager': 'DM',
        'Cleaner': 'CL',
        'Junior Lifeguard': 'JL',
        'Training': 'T',
        'Pool Shop QLD': 'PS'
    }

    return role_mapping.get(role, role)

 

def apply_borders_to_cells(ws):
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
   

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.border = thin_border


def bold_headers_and_roles(ws):
    # Bolding the headers
    for cell in ws["A2":"F2"][0]:
        cell.font = Font(bold=True)
  
    # Bolding the roles column
    for cell in ws["A"]:
        cell.font = Font(bold=True)
 
 
def apply_thick_border_above_first_occurrence(ws, value_to_check):
    # Define the thick border style
    thick_border_top = Border(top=Side(border_style="thick", color="000000"))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            # Check if value_to_check is a substring of cell.value
            if value_to_check in str(cell.value):
                for cell_in_row in row:
                    current_border = cell_in_row.border
                    new_border = Border(top=thick_border_top.top,
                                        left=current_border.left,
                                        right=current_border.right,
                                        bottom=current_border.bottom)
                    cell_in_row.border = new_border
                return  # Exit once we applied the border


 

 

def apply_thick_border_on_last_occurrence(ws, value_to_check):
    # Define the thick border style
    thick_border_bottom = Border(bottom=Side(border_style="thick", color="000000"))
    
    last_found_row = None
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            if cell.value == value_to_check:
                last_found_row = cell.row

    if last_found_row:
        for col in range(1, 7):  # Adjusting for all columns A to F
            cell = ws.cell(row=last_found_row, column=col)
            current_border = cell.border
            new_border = Border(bottom=thick_border_bottom.bottom,
                                left=current_border.left,
                                right=current_border.right,
                                top=current_border.top)
            cell.border = new_border



 

def alternate_row_fill(ws):
    light_grey_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    last_start_time = None
    fill_flag = False


    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=6):
        # Extract start time
        shift_time = row[2].value
        current_start_time = shift_time.split("-")[0].strip()

        if last_start_time and current_start_time != last_start_time:
            fill_flag = not fill_flag
 
        if fill_flag:
            for cell in row:
                cell.fill = light_grey_fill
 
        last_start_time = current_start_time
 
def set_column_widths(ws):
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 19
    ws.column_dimensions['D'].width = 13
    ws.column_dimensions['E'].width = 19
    ws.column_dimensions['F'].width = 20
 
 
 
def populate_runsheet(ws, col, day_name, use_initials=False):
    date_value = get_date_from_day_name(day_name)
    ws.merge_cells('A1:F1')
    ws['A1'] = date_value
    headers = ["Role", "Name", "Shift Time", "Breaks", "Actual Hours", "Job"]
    ws.append(headers)
 
    shifts_for_person_role = {}
    for row in range(2, sheet_roster.max_row + 1):
        name_cell_value = sheet_roster.cell(row=row, column=1).value
        name = name_cell_value.split()[0] if name_cell_value else ""
        cell_value = sheet_roster.cell(row=row, column=col).value
        if cell_value and "-" in cell_value:
            roles_and_times = extract_role_and_time(cell_value)
 
            for abbreviated_role, time_string in roles_and_times:
                if time_string:
                    start_time, finish_time = [parse_time(x.strip()) for x in time_string.split("-")]
                    shift_time = f"{start_time.strftime('%H:%M')} - {finish_time.strftime('%H:%M')}"
                    breaks = calculate_breaks(start_time, finish_time)
                    actual_hours = ""
                    job = ""
 
                    key = (name, abbreviated_role, shift_time)
                    shifts_for_person_role[key] = [abbreviated_role, name, shift_time, breaks, "", actual_hours]
 
    rows_for_day = list(shifts_for_person_role.values())
    sorted_rows = sorted(rows_for_day, key=sort_key)
    for sorted_row in sorted_rows:
        ws.append(sorted_row)
 
    # Formatting
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = centered_alignment
    apply_borders_to_cells(ws)
    bold_headers_and_roles(ws)
    apply_thick_border_above_first_occurrence(ws, "DM")
    apply_thick_border_above_first_occurrence(ws, "CL")
    apply_thick_border_on_last_occurrence(ws_full, "CL")

    alternate_row_fill(ws)
    set_column_widths(ws)   

 
current_directory = os.path.dirname(os.path.abspath(__file__))

root = tk.Tk()
root.withdraw()  # Hide the root window
roster_file_path = filedialog.askopenfilename(title="Select the Roster File", filetypes=[("Excel Files", "*.xls;*.xlsx")])

if not roster_file_path:
    raise FileNotFoundError("No file was selected")

print("Selected file:", roster_file_path)  # Debug print to confirm file selection

# Convert .xls to .xlsx if necessary
if roster_file_path.endswith('.xls'):
    df = pd.read_excel(roster_file_path)
    new_file_path = roster_file_path[:-4] + ".xlsx"  # Name for the converted file
    df.to_excel(new_file_path, index=False)  # Save as .xlsx
    os.remove(roster_file_path)  # Remove the original .xls file
    roster_file_path = new_file_path  # Update the path to use the new .xlsx file


 


wb_runsheet = openpyxl.Workbook()
deafult_sheet =wb_runsheet.active
wb_runsheet.remove(deafult_sheet)
wb_roster = openpyxl.load_workbook(roster_file_path)
 
 
sheet_roster = wb_roster.active
 
 
day_columns = {
    2: "Monday",
    3: "Tuesday",
    4: "Wednesday",
    5: "Thursday",
    6: "Friday",
    7: "Saturday",
    8: "Sunday"
}
 
centered_alignment = Alignment(horizontal="center")
 
for col, day_name in day_columns.items():
    # For full names
    ws_full = wb_runsheet.create_sheet(f"{day_name}")
    populate_runsheet(ws_full, col, day_name, use_initials=False)
 


         
 
save_path = os.path.join(current_directory, "runsheet_by_day.xlsx")

wb_runsheet.save(save_path)